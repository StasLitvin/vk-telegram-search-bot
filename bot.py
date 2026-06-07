import asyncio
import os
import re
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, \
    ReplyKeyboardRemove
from aiogram.exceptions import TelegramBadRequest, TelegramRetryAfter
from dotenv import load_dotenv
from vk_parser import VKParser
from tg_parser import TGParser
import openpyxl
from datetime import datetime, timedelta
from screenshot_helper import ScreenshotHelper
from openpyxl.drawing.image import Image as ExcelImage
from pathlib import Path
import shutil

load_dotenv()

bot = Bot(token=os.getenv('BOT_TOKEN'))
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

class ParserStates(StatesGroup):
    waiting_for_links = State()
    waiting_for_keywords = State()
    parsing = State()

def get_main_menu():
    """Главное меню"""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Начать парсинг")],
            [KeyboardButton(text="Инструкция"), KeyboardButton(text="О боте")]
        ],
        resize_keyboard=True
    )
    return keyboard

def get_cancel_keyboard():
    """Клавиатура с кнопкой отмены"""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Отменить")]
        ],
        resize_keyboard=True
    )
    return keyboard

def get_confirm_keyboard():
    """Inline-клавиатура подтверждения"""
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Начать парсинг", callback_data="confirm_parse"),
                InlineKeyboardButton(text="Изменить ключевые слова", callback_data="change_keywords")
            ],
            [InlineKeyboardButton(text="Отменить", callback_data="cancel")]
        ]
    )
    return keyboard

def get_restart_keyboard():
    """Клавиатура после завершения"""
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Новый парсинг", callback_data="restart")],
            [InlineKeyboardButton(text="Главное меню", callback_data="main_menu")]
        ]
    )
    return keyboard

def extract_links_from_text(text):
    """
    Извлечение всех ссылок из текста

    Поддерживаемые форматы:
    - https://vk.com/rosmolodez
    - vk.com/public123
    - t.me/channelname
    - @channelname (Telegram)
    """
    links = []

    https_pattern = r'https?://[^\s\)\],<>"\']+'
    https_matches = re.findall(https_pattern, text)
    links.extend(https_matches)

    text_cleaned = text
    for match in https_matches:
        text_cleaned = text_cleaned.replace(match, '')

    vk_pattern = r'\bvk\.com/([a-zA-Z0-9_]+)'
    vk_matches = re.findall(vk_pattern, text_cleaned)
    for match in vk_matches:
        full_link = f"https://vk.com/{match}"
        if full_link not in links:
            links.append(full_link)

    tg_pattern = r'\bt\.me/([a-zA-Z0-9_]+)'
    tg_matches = re.findall(tg_pattern, text_cleaned)
    for match in tg_matches:
        full_link = f"https://t.me/{match}"
        if full_link not in links:
            links.append(full_link)

    tg_username_pattern = r'@([a-zA-Z0-9_]{5,})'
    tg_username_matches = re.findall(tg_username_pattern, text)
    for match in tg_username_matches:
        tg_link = f"https://t.me/{match}"
        if tg_link not in links:
            links.append(tg_link)

    return list(set(links))

def parse_keywords(text):
    """Парсинг ключевых слов с извлечением ссылок"""
    criteria = []

    lines = text.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        extracted_links = extract_links_from_text(line)

        clean_line = line
        for link in extracted_links:
            clean_line = re.sub(r'https?://[^\s\)\],<>"\']+', '', clean_line)
            clean_line = re.sub(r'vk\.com/[^\s\)\],<>"\']+', '', clean_line)
            clean_line = re.sub(r't\.me/[^\s\)\],<>"\']+', '', clean_line)
            clean_line = re.sub(r'@[a-zA-Z0-9_]{5,}', '', clean_line)
            clean_line = re.sub(r'club\d+', '', clean_line)
            clean_line = re.sub(r'public\d+', '', clean_line)

        clean_line = clean_line.strip()
        clean_line = re.sub(r'^\+\s*|\s*\+$', '', clean_line).strip()
        clean_line = re.sub(r'\s*\+\s*', ' + ', clean_line)

        clean_line = clean_line.replace('"', '').replace('"', '').replace('"', '')
        clean_line = clean_line.replace("'", '').replace("'", '').replace("'", '')

        text_elements = []

        if clean_line:
            segments = clean_line.split('+')
            for segment in segments:
                segment = segment.strip()
                if segment:
                    text_elements.append(segment)

        if text_elements or extracted_links:
            criterion = {
                'text_elements': text_elements,
                'link_elements': extracted_links,
                'type': 'AND'
            }
            criteria.append(criterion)

            print(f"DEBUG: Создан критерий:")
            print(f"  Текст: {text_elements}")
            print(f"  Ссылки: {extracted_links}")

    return {
        'criteria': criteria
    }

def get_progress_bar(current, total, length=10):
    """Создание визуального прогресс-бара"""
    filled = int(length * current / total)
    bar = '' * filled + '' * (length - filled)
    return f"{bar} {int(current / total * 100)}%"

async def safe_edit_message(msg, text, parse_mode="Markdown", max_retries=3):
    """Безопасное редактирование с повторными попытками"""
    for attempt in range(max_retries):
        try:
            await msg.edit_text(text, parse_mode=parse_mode)
            return True
        except TelegramRetryAfter as e:
            print(f"Флуд-контроль, ждём {e.retry_after} сек...")
            await asyncio.sleep(e.retry_after)
        except TelegramBadRequest as e:
            error_str = str(e).lower()
            if "message is not modified" in error_str:
                return True
            elif "no text in the message" in error_str:
                print("Попытка редактировать медиа-сообщение")
                return False
            elif "message can't be edited" in error_str:
                print("Сообщение нельзя редактировать")
                return False
            print(f"Ошибка редактирования: {e}")
            return False
        except Exception as e:
            print(f"Ошибка редактирования (попытка {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                await asyncio.sleep(2)
    return False

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "*Добро пожаловать!*\n\n"
        "Я помогу найти посты в VK и Telegram по ключевым словам и ссылкам.\n\n"
        "Выберите действие:",
        parse_mode="Markdown",
        reply_markup=get_main_menu()
    )

@dp.message(F.text == "Начать парсинг")
async def start_parsing_flow(message: types.Message, state: FSMContext):
    await message.answer(
        "*Шаг 1/2: Ссылки на источники*\n\n"
        "Отправьте список ссылок на группы ВКонтакте или каналы Telegram "
        "(по одной на строку):\n\n"
        "*Примеры:*\n"
        "```\n"
        "https://vk.com/public123456\n"
        "https://t.me/channelname\n"
        "@channelname\n"
        "```",
        parse_mode="Markdown",
        reply_markup=get_cancel_keyboard()
    )
    await state.set_state(ParserStates.waiting_for_links)

@dp.message(F.text == "Инструкция")
async def show_instructions(message: types.Message):
    await message.answer(
        "*Инструкция по использованию*\n\n"
        "*1Типы поиска:*\n\n"
        "*OR-поиск* (одно из слов):\n"
        "```\n"
        "скидка\n"
        "акция\n"
        "распродажа\n"
        "```\n"
        "_Найдёт посты, где есть хотя бы одно слово_\n\n"
        "*AND-поиск* (все слова обязательно):\n"
        "```\n"
        "скидка + акция\n"
        "новинка + 2024 + купить\n"
        "```\n"
        "_Найдёт только посты со ВСЕМИ словами_\n\n"
        "*Поиск по ссылкам* (OR-логика):\n"
        "```\n"
        "https://vk.com/rosmolodez\n"
        "https://t.me/channelname\n"
        "```\n\n"
        "*Комбинированный поиск* (текст + ссылки):\n"
        "```\n"
        "\"программа развития\" + Росмолодёжь + https://vk.com/rosmolodez.grants\n"
        "```\n"
        "_Найдёт посты, где есть И текст, И ссылка_\n\n"
        "*2Лимиты:*\n"
        "• VK: до 300 последних постов\n"
        "• Telegram: до 1000 последних постов\n\n"
        "*3Результат:*\n"
        "Excel-файл со ссылками, датами, просмотрами и текстом постов",
        parse_mode="Markdown",
        reply_markup=get_main_menu()
    )

@dp.message(F.text == "О боте")
async def show_about(message: types.Message):
    await message.answer(
        "*О боте*\n\n"
        "*Парсер постов VK & Telegram*\n"
        "Версия: 2.1\n\n"
        "*Возможности:*\n"
        "Парсинг VK и Telegram\n"
        "Поиск по ключевым словам\n"
        "Поиск по ссылкам в постах\n"
        "AND/OR логика поиска\n"
        "Комбинированный поиск (текст + ссылки)\n"
        "Экспорт в Excel\n"
        "Поддержка множества источников\n\n"
        "*Новое в 2.1:*\n"
        "Поиск постов по упоминанию конкретных ссылок\n"
        "Комбинированные запросы (текст И ссылка)",
        parse_mode="Markdown",
        reply_markup=get_main_menu()
    )

@dp.message(F.text == "Отменить")
async def cancel_operation(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "Операция отменена.\n\n"
        "Возвращаю в главное меню.",
        reply_markup=get_main_menu()
    )

@dp.message(ParserStates.waiting_for_links)
async def process_links(message: types.Message, state: FSMContext):
    if message.text == "Отменить":
        await cancel_operation(message, state)
        return

    links = [link.strip() for link in message.text.split('\n') if link.strip()]

    if not links:
        await message.answer(
            "Не найдено ни одной ссылки. Попробуйте еще раз.",
            reply_markup=get_cancel_keyboard()
        )
        return

    await state.update_data(links=links)

    await message.answer(
        f"Получено *{len(links)}* ссылок\n\n"
        f"*Шаг 2/2: Ключевые слова и ссылки*\n\n"
        f"Отправьте ключевые слова или ссылки для поиска:\n\n"
        f"*Варианты:*\n\n"
        f"1*Простой поиск:*\n"
        f"```\n"
        f"скидка\n"
        f"акция\n"
        f"```\n\n"
        f"2*Поиск с AND (+):*\n"
        f"```\n"
        f"скидка + акция\n"
        f"```\n\n"
        f"3*Поиск по ссылкам:*\n"
        f"```\n"
        f"https://vk.com/rosmolodez\n"
        f"https://t.me/channelname\n"
        f"```\n\n"
        f"4*Комбинированный поиск:*\n"
        f"```\n"
        f"\"программа\" + Росмолодёжь + https://vk.com/rosmolodez.grants\n"
        f"```",
        parse_mode="Markdown",
        reply_markup=get_cancel_keyboard()
    )
    await state.set_state(ParserStates.waiting_for_keywords)

@dp.message(ParserStates.waiting_for_keywords)
async def process_keywords(message: types.Message, state: FSMContext):
    if message.text == "Отменить":
        await cancel_operation(message, state)
        return

    keywords_data = parse_keywords(message.text)
    criteria = keywords_data.get('criteria', [])

    if not criteria:
        await message.answer(
            "Не найдено ни одного критерия поиска. Попробуйте еще раз.",
            reply_markup=get_cancel_keyboard()
        )
        return

    preview_lines = []

    for i, criterion in enumerate(criteria[:5], 1):
        text_elems = criterion.get('text_elements', [])
        link_elems = criterion.get('link_elements', [])

        parts = []
        if text_elems:
            parts.extend([f"`{t[:30]}{'...' if len(t) > 30 else ''}`" for t in text_elems])
        if link_elems:
            parts.extend([f"`{l[:35]}...`" for l in link_elems])

        preview_lines.append(f"{i}. {' **+** '.join(parts)} *(AND)*")

    if len(criteria) > 5:
        preview_lines.append(f"_...и ещё {len(criteria) - 5} критериев_")

    criteria_preview = "\n".join(preview_lines)

    await state.update_data(keywords=keywords_data)
    data = await state.get_data()

    await message.answer(
        f"*Готово к запуску!*\n\n"
        f"*Критерии поиска ({len(criteria)}):*\n"
        f"{criteria_preview}\n\n"
        f"*Источников:* {len(data['links'])}\n\n"
        f"*Логика поиска:*\n"
        f"Внутри каждого критерия: **AND** (все элементы обязательны)\n"
        f"Между критериями: **OR** (достаточно одного)\n\n"
        f"Подтвердите запуск:",
        parse_mode="Markdown",
        reply_markup=get_confirm_keyboard()
    )
    await state.set_state(ParserStates.parsing)

@dp.callback_query(F.data == "confirm_parse")
async def confirm_parsing(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()

    data = await state.get_data()
    criteria = data['keywords'].get('criteria', [])
    total_elements = 0
    for criterion in criteria:
        total_elements += len(criterion.get('text_elements', []))
        total_elements += len(criterion.get('link_elements', []))

    status_msg = await callback.message.answer(
        f"*Начинаю парсинг...*\n\n"
        f"Источников: {len(data['links'])}\n"
        f"Критериев поиска: {len(criteria)}\n"
        f"Всего элементов: {total_elements}\n\n"
        f"Подождите, это может занять несколько минут...",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    results = await parse_all(data['links'], data['keywords'], callback.message, status_msg)

    if not results:
        await callback.message.answer(
            "*Посты не найдены*\n\n"
            "Попробуйте изменить критерии поиска или источники.",
            parse_mode="Markdown",
            reply_markup=get_restart_keyboard()
        )
        try:
            await status_msg.delete()
        except:
            pass
        await state.clear()
        return

    try:

        result = await create_excel_with_screenshots(results, status_msg, compress_images=True)

        if isinstance(result, tuple):
            file_path, file_size_mb = result
        else:
            file_path = result
            file_size_mb = Path(file_path).stat().st_size / (1024 * 1024)

    except Exception as e:
        print(f"Ошибка создания скриншотов: {e}")
        import traceback
        traceback.print_exc()

        await safe_edit_message(
            status_msg,
            f"*Ошибка создания скриншотов*\n\n"
            f"Создаю отчёт без изображений..."
        )
        file_path = create_excel(results)
        file_size_mb = Path(file_path).stat().st_size / (1024 * 1024)

    TELEGRAM_LIMIT_MB = 50

    if file_size_mb > TELEGRAM_LIMIT_MB:
        print(f"Файл слишком большой: {file_size_mb:.2f} МБ")

        await safe_edit_message(
            status_msg,
            f"*Файл слишком большой*\n\n"
            f"Размер: {file_size_mb:.1f} МБ (лимит Telegram: 50 МБ)\n\n"
            f"Создаю облегчённую версию без скриншотов..."
        )

        try:
            os.remove(file_path)
        except:
            pass

        file_path = create_excel(results)
        file_size_mb = Path(file_path).stat().st_size / (1024 * 1024)

        file = types.FSInputFile(file_path)

        caption_text = (
            f"*Парсинг завершён!*\n\n"
            f"Найдено: *{len(results)}* постов\n"
            f"Источников: {len(data['links'])}\n"
            f"Критериев поиска: {len(criteria)}\n\n"
            f"Версия без скриншотов (файл был > 50 МБ)\n"
            f"Размер: {file_size_mb:.1f} МБ"
        )

    else:
        file = types.FSInputFile(file_path)

        caption_text = (
            f"*Парсинг завершён!*\n\n"
            f"Найдено: *{len(results)}* постов\n"
            f"Источников: {len(data['links'])}\n"
            f"Критериев поиска: {len(criteria)}\n"
            f"Размер файла: {file_size_mb:.1f} МБ"
        )

        if 'compressed' in file_path:
            caption_text += f"\nСкриншоты: сжаты для уменьшения размера"

    try:
        await callback.message.answer_document(
            file,
            caption=caption_text,
            parse_mode="Markdown"
        )
    except Exception as e:

        print(f"Не удалось отправить файл: {e}")

        try:
            os.remove(file_path)
        except:
            pass

        file_path = create_excel(results)
        file = types.FSInputFile(file_path)

        await callback.message.answer_document(
            file,
            caption=f"*Парсинг завершён!*\n\n"
                    f"Найдено: *{len(results)}* постов\n"
                    f"Версия без скриншотов (файл был слишком большой)",
            parse_mode="Markdown"
        )

    await callback.message.answer(
        "Выберите действие:",
        reply_markup=get_restart_keyboard()
    )

    os.remove(file_path)
    try:
        await status_msg.delete()
    except:
        pass
    await state.clear()

@dp.callback_query(F.data == "change_keywords")
async def change_keywords(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    await callback.message.edit_text(
        "*Изменение ключевых слов*\n\n"
        "Отправьте новые ключевые слова:",
        parse_mode="Markdown"
    )
    await callback.message.answer(
        "Введите новые ключевые слова:",
        reply_markup=get_cancel_keyboard()
    )
    await state.set_state(ParserStates.waiting_for_keywords)

@dp.callback_query(F.data == "cancel")
async def cancel_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer("Операция отменена")
    await state.clear()
    await callback.message.edit_text(
        "Операция отменена."
    )
    await callback.message.answer(
        "Возвращаю в главное меню.",
        reply_markup=get_main_menu()
    )

@dp.callback_query(F.data == "restart")
async def restart_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer("Перезапуск...")
    await state.clear()

    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except:
        pass

    new_message = await callback.message.answer(
        "Перезапускаю бота...",
        reply_markup=ReplyKeyboardRemove()
    )

    await start_parsing_flow(new_message, state)

@dp.callback_query(F.data == "main_menu")
async def main_menu_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer("Главное меню")
    await state.clear()

    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except:
        pass

    await callback.message.answer(
        "Выберите действие:",
        reply_markup=get_main_menu()
    )

async def parse_all(links, keywords, message, status_msg):
    results = []
    vk_parser = VKParser(os.getenv('VK_TOKEN'))
    tg_parser = TGParser(os.getenv('TG_API_ID'), os.getenv('TG_API_HASH'))

    errors = []
    last_update_time = datetime.now()
    min_update_interval = timedelta(seconds=2)

    try:
        await tg_parser.start()

        for i, link in enumerate(links, 1):
            try:
                current_time = datetime.now()
                if i == 1 or i == len(links) or current_time - last_update_time >= min_update_interval:
                    progress_bar = get_progress_bar(i, len(links))

                    success = await safe_edit_message(
                        status_msg,
                        f"*Парсинг источников*\n\n"
                        f"{progress_bar}\n"
                        f"*{i}/{len(links)}* источников\n\n"
                        f"`{link[:35]}...`\n\n"
                        f"Найдено: *{len(results)}* постов"
                    )

                    if success:
                        last_update_time = current_time

                if 'vk.com' in link:
                    posts = await vk_parser.parse_group(link, keywords)

                    for post in posts:
                        post['platform'] = 'vk'
                    results.extend(posts)
                elif 't.me' in link or link.startswith('@'):
                    posts = await tg_parser.parse_channel(link, keywords)
                    for post in posts:
                        post['platform'] = 'telegram'
                    results.extend(posts)

                await asyncio.sleep(0.3)

            except Exception as e:
                error_text = f"{link}: {str(e)[:50]}"
                errors.append(error_text)
                print(f"{error_text}")

        await tg_parser.stop()

        if errors:
            errors_text = "\n".join([f"• `{err[:60]}`" for err in errors[:3]])
            if len(errors) > 3:
                errors_text += f"\n_...и ещё {len(errors) - 3}_"

            await message.answer(
                f"*Предупреждения*\n\n"
                f"Ошибки при обработке {len(errors)} источников:\n"
                f"{errors_text}",
                parse_mode="Markdown"
            )

    except Exception as e:
        print(f"Критическая ошибка: {e}")

    return results

async def create_excel_with_screenshots(results, status_msg, compress_images=True):
    """
    Создание Excel с автоматическим сжатием и проверкой размера

    Args:
        compress_images: если True, сжимает изображения для уменьшения размера файла
    """
    try:
        from PIL import Image as PILImage
        print("Pillow доступен")
    except ImportError:
        print("Pillow не установлен!")
        await safe_edit_message(
            status_msg,
            f"*Ошибка: Pillow не установлен*\n\n"
            f"Создаю отчёт без изображений..."
        )
        return create_excel(results)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    screenshot_helper = ScreenshotHelper()
    screenshots_enabled = await screenshot_helper.check_availability()

    if not screenshots_enabled:
        await safe_edit_message(
            status_msg,
            f"*Скриншоты недоступны*\n\n"
            f"Создаю отчёт без изображений..."
        )
        return create_excel(results)

    headers = ['Источник', 'Ссылка на источник', 'Ссылка на пост', 'Дата', 'Просмотры', 'Найденные ссылки', 'Текст',
               'Скриншот']
    ws.append(headers)

    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for idx, result in enumerate(results):
        result['index'] = idx

    if len(results) > 100:
        max_screenshots = 200
    elif len(results) > 50:
        max_screenshots = 200
    else:
        max_screenshots = 200

    screenshot_tasks = []
    for result in results[:max_screenshots]:
        screenshot_tasks.append({
            'index': result['index'],
            'link': result['link'],
            'platform': result['platform']
        })

    vk_tasks = [t for t in screenshot_tasks if t['platform'] == 'vk']
    tg_tasks = [t for t in screenshot_tasks if t['platform'] == 'telegram']

    await safe_edit_message(
        status_msg,
        f"*Создание скриншотов*\n\n"
        f"Обрабатываю {len(screenshot_tasks)} из {len(results)} постов:\n"
        f"   • Telegram: {len(tg_tasks)}\n"
        f"   • VK: {len(vk_tasks)}\n\n"
        f"Примерное время: {len(screenshot_tasks) * 6 // 60 + 1} мин"
    )

    await screenshot_helper.start()

    if not screenshot_helper.is_available:
        await screenshot_helper.stop()
        await safe_edit_message(
            status_msg,
            f"*Не удалось запустить браузер*\n\n"
            f"Создаю отчёт без скриншотов..."
        )
        return create_excel(results)

    screenshots = await screenshot_helper.take_screenshots_batch(
        screenshot_tasks,
        max_concurrent=1,
        enable_vk=True
    )

    await screenshot_helper.stop()

    print(f"\nDEBUG: Всего screenshot_tasks: {len(screenshot_tasks)}")
    print(f"DEBUG: Первые 3 задачи: {screenshot_tasks[:3]}")
    print(f"DEBUG: Ключи screenshots dict: {sorted(screenshots.keys())[:10]}")
    print(f"DEBUG: Индекс 0 в screenshots? {0 in screenshots}")
    if 0 in screenshots:
        print(f"DEBUG: Путь скриншота 0: {screenshots[0]}")
        from pathlib import Path as P
        if P(screenshots[0]).exists():
            print(f"DEBUG: Размер файла 0: {P(screenshots[0]).stat().st_size} байт")
        else:
            print(f"DEBUG: ФАЙЛ НЕ СУЩЕСТВУЕТ!")
    else:
        print(f"DEBUG: ИНДЕКС 0 ОТСУТСТВУЕТ В РЕЗУЛЬТАТАХ!")

    screenshots_count = len(screenshots)
    await safe_edit_message(
        status_msg,
        f"Создано {screenshots_count} скриншотов\n\n"
        f"Формирую Excel..."
    )

    row_height = 240

    for idx, result in enumerate(results, start=2):
        date_formatted = result['date'].strftime('%d.%m.%Y %H:%M')
        found_links_text = '\n'.join(result.get('found_links', [])) if result.get('found_links') else ''

        ws.append([
            result['source_name'],
            result['source_link'],
            result['link'],
            date_formatted,
            result['views'],
            found_links_text,
            result['text'][:150] + '...' if len(result['text']) > 150 else result['text'],
            ''
        ])

        result_index = result['index']
        if result_index in screenshots:
            ws.row_dimensions[idx].height = row_height

            img_path = screenshots[result_index]

            try:
                from pathlib import Path
                if not Path(img_path).exists():
                    print(f"Файл не найден: {img_path}")
                    continue

                file_size = Path(img_path).stat().st_size
                if file_size < 100:
                    print(f"Файл слишком маленький: {img_path}")
                    continue

                pil_img = PILImage.open(img_path)

                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                    if pil_img.mode == 'P':
                        pil_img = pil_img.convert('RGBA')
                    background.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                    pil_img = background

                if compress_images:

                    new_width = int(pil_img.width * 0.5)
                    new_height = int(pil_img.height * 0.5)
                    pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)

                optimized_path = img_path.replace('.png', '_opt.jpg')

                pil_img.save(optimized_path, 'JPEG', quality=85, optimize=True)

                excel_img = openpyxl.drawing.image.Image(optimized_path)

                excel_img.width = 180
                excel_img.height = 220
                excel_img.anchor = f'H{idx}'
                ws.add_image(excel_img)

            except Exception as e:
                print(f"Не удалось вставить {img_path}: {e}")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 30

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(
                horizontal='left',
                vertical='top',
                wrap_text=True
            )

    for row_idx in range(2, ws.max_row + 1):
        source_link_cell = ws.cell(row=row_idx, column=2)
        if source_link_cell.value:
            source_link_cell.hyperlink = source_link_cell.value
            source_link_cell.style = "Hyperlink"

        post_link_cell = ws.cell(row=row_idx, column=3)
        if post_link_cell.value:
            post_link_cell.hyperlink = post_link_cell.value
            post_link_cell.style = "Hyperlink"

    filename = f'results_compressed_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)

    file_size_mb = Path(filename).stat().st_size / (1024 * 1024)
    print(f"\nРазмер файла: {file_size_mb:.2f} МБ")

    try:
        shutil.rmtree('screenshots')
    except Exception as e:
        print(f"Не удалось удалить temp: {e}")

    return filename, file_size_mb

def create_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    headers = ['Источник', 'Ссылка на источник', 'Ссылка на пост', 'Дата', 'Просмотры', 'Найденные ссылки', 'Текст']
    ws.append(headers)

    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for result in results:
        date_formatted = result['date'].strftime('%d.%m.%Y %H:%M')

        found_links_text = '\n'.join(result.get('found_links', [])) if result.get('found_links') else ''

        ws.append([
            result['source_name'],
            result['source_link'],
            result['link'],
            date_formatted,
            result['views'],
            found_links_text,
            result['text'][:100] + '...' if len(result['text']) > 100 else result['text']
        ])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for row_idx in range(2, ws.max_row + 1):
        source_link_cell = ws.cell(row=row_idx, column=2)
        source_link_cell.hyperlink = source_link_cell.value
        source_link_cell.style = "Hyperlink"

        post_link_cell = ws.cell(row=row_idx, column=3)
        post_link_cell.hyperlink = post_link_cell.value
        post_link_cell.style = "Hyperlink"

    filename = f'results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    return filename

async def main():
    print("Бот запущен!")
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
